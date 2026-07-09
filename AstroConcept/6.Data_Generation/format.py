from openai import OpenAI
import openpyxl
client = OpenAI(
    base_url='https://api.openai-proxy.org/v1',
    api_key='sk-J2vxk7sWso4VvYQfDo20h9WY4Fd5Q4dnyDAgNLJSfiJE4vOl',
)

wb = openpyxl.load_workbook(r"D:\PythonProjects\AstroCalcBoost\main\AstroConcept\7.概念图构建蒸馏数据\steps7_1-1107-answer思维链蒸馏.xlsx")
sheet = wb.active  # 假设数据在第一个工作表
# 创建一个列表来存储结果
results = []

for i, row in enumerate(sheet.iter_rows(min_row=2)):  # 假设第一行是标题行，从第二行开始读取
        # 读取每行中特定列的值
    # try:
        question = row[7].value.strip()  # 假设question在第二列
        answer = row[4].value.strip()
        # total = row[4].value
        prompt = (r"""Please process the given text content according to the following rules:
        1. Format the final answer(s) using LaTeX's \boxed{} command:   
        2. Box each answer separately:  
           - Keep all math notation (LaTeX, units, vectors) unchanged.  
           - Use `\text{}` for units (e.g., `\boxed{3.2\ \text{m/s}}`).  
        3. Do not modify any intermediate steps or non-answer text. """
                  f'Text to Process:  {question}')
        print(prompt)
        # 将答案添加到结果列表中
        for id, j in zip(range(1),range(1)):
            try:
                response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        # {"role": "system", "content": "You are a helpful assistant."},
                        {"role": "user",
                         "content": str(prompt)}
                    ],
                    stream=False,
                    # temperature=0
                )
                print(response.choices[0].message.content)
                sheet.cell(row=i + 2, column=j + 10, value=response.choices[0].message.content)  # 第十列的索引是11（因为索引是从1开始的）
            except Exception as e:
                print("无法回答")
                sheet.cell(row=i + 2, column=j + 10, value="无法回答")
        wb.save(r"D:\PythonProjects\AstroCalcBoost\main\AstroConcept\9. 答案格式化\steps7_1-1107-answer思维链蒸馏.xlsx")
        print(f"第{i}行已生成完毕！")
        print('-' * 100)  # 分隔线
