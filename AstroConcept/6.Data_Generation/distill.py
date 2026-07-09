# -*- coding: utf-8 -*-
#  道阻且长，行则将至
# -----Sunnyln---
#  1/9/2024  12:07 PM
from openai import OpenAI
import openpyxl

client = OpenAI(api_key='sk-6044b09e09ea4011a830fbf0f9cbf383',base_url="https://api.deepseek.com")

wb = openpyxl.load_workbook(r"D:\PythonProjects\AstroCalcBoost\main\AstroConcept\6.数据生成\output_walk_steps_8\Planets_output_walk_result_steps_8.xlsx")
sheet = wb.active  # 假设数据在第一个工作表
# 创建一个列表来存储结果
results = []
# sheet.cell(row=1, column=8, value="推理内容")  # 第10列的列名
# sheet.cell(row=1, column=9, value="回答内容")  # 第11列的列名


for i, row in enumerate(sheet.iter_rows(min_row=2)):  # 假设第一行是标题行，从第二行开始读取
        # 读取每行中特定列的值
    # try:
        question = row[4].value.strip()  # 假设question在第二列
        answer = row[5].value.strip()
        prompt = (
            f'A conversation between User and Assistant. The user asks a question, and the Assistant solves it.\n '
            f'The assistant first thinks about the reasoning process in the mind and then provides the user with the answer, when necessary referring to the standard answer.\n'
            f'The reasoning process and answer are enclosed within <think> </think> and <answer> </answer> tags, respectively, i.e., <think> reasoning process here </think: <answer> answer here </answer>.\n '
            f'Question: {question}\n'
            f'Reference answer: {answer}\n')

        print(prompt)
        print("\n" + "-" * 50)
        # 将答案添加到结果列表中
        for id, j in zip(range(1),range(1)):
            try:
                response = client.chat.completions.create(
                    model="deepseek-reasoner",
                    messages=[
                        # {"role": "system", "content": "You are a helpful assistant."},
                        {"role": "user",
                         "content": str(prompt)}
                    ],
                    stream=False,
                    # temperature=0
                )
                reasoning_content = response.choices[0].message.reasoning_content
                answer_content = response.choices[0].message.content
                print(f'第{i}题推理内容为：{reasoning_content}\n')
                print(f'第{i}题回答内容为：{answer_content}')
                # print("\n" + "-" * 100)
                sheet.cell(row=i + 2, column=id + 7, value=reasoning_content)  # 第十列的索引是11（因为索引是从1开始的）
                sheet.cell(row=i + 2, column=id + 8, value=answer_content)
                # sheet.cell(row=i + 2, column=j +7, value=response.choices[0].message.content)  # 第十列的索引是11（因为索引是从1开始的）
            except Exception as e:
                print({str(e)})
                sheet.cell(row=i + 2, column=j + 7, value="无法回答")
        wb.save(r"D:\PythonProjects\AstroCalcBoost\main\AstroConcept\7.概念图构建蒸馏数据\output_walk_steps_8\Planets_output_walk_result_steps_8.xlsx")
        print(f'第{i}行已回答完毕！')
        print("\n" + "-" * 100)

