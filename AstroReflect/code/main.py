import re
import openpyxl
from openai import OpenAI
import concurrent.futures
from processor import extract_calculations_with_equals  # 可选保留
from formula_retrieval import AstronomyKnowledgeEngine
from dimensional_verifier import DimensionalVerifier
import json
from code_executor import PythonSession
from majority_voting import correctness_reward_func, get_majority_vote_answer
# 配置
client = OpenAI(api_key="sk-d688345650f045369075d1839492c9ba", base_url="https://api.deepseek.com")
FORMULA_PATH = r"D:\PythonProjects\AstroCalcBoost2\main\formula\astronomy_formulas_overall.jsonl"
formula_retrieval_engine = AstronomyKnowledgeEngine(FORMULA_PATH)
# code_executor = AstroCodeExecutor()


class DecompositionAgent:
    """
    规划器：负责将复杂的天文物理问题拆解为多个核心知识点查询
    """
    def __init__(self, client):
        self.client = client

    def decompose(self, question):
        """
        输入: "如果太阳变成同质量黑洞，地球轨道处的引力红移是多少？"
        输出: ["Schwarzschild Radius formula", "Gravitational Redshift formula", "Solar Mass constant"]
        """
        # System Prompt：强制要求输出 JSON 格式的列表
        system_prompt = (
            "You are a research planner for an astronomy problem solver.\n"
            "Your task: Break down the user's complex problem into 1-2 distinct, specific search queries to find the necessary astronomical formulas.\n"
            "Rules:\n"
            "1. Output ONLY a raw JSON list of strings. No markdown formatting, no explanations.\n"
            "2. Each string should be a standard physics/astronomy term (e.g., 'Kepler's Third Law', 'Distance Modulus').\n"
            "3. Do not try to solve the problem, just plan the search.\n"
        )
        print(f'系统提示词：{system_prompt}')
        print("===================================================================================")

        user_prompt = f"Problem: \"{question}\"\nSearch Queries:"
        print(f'用户提问：{user_prompt}')
        print("===================================================================================")

        try:
            response = self.client.chat.completions.create(
                model="deepseek-chat", # 使用小模型即可，速度快
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                # temperature=0.1 # 必须低温，保证格式稳定
            )
            content = response.choices[0].message.content.strip()

            # --- 鲁棒性处理：清洗可能的 Markdown 标记 ---
            # 有时候模型会手欠加上 ```json ... ```，这里去掉它
            if "```" in content:
                content = re.sub(r"```(?:json)?(.*?)```", r"\1", content, flags=re.DOTALL).strip()

            queries = json.loads(content)

            # 确保返回的是列表
            if isinstance(queries, list):
                print(f"🤔 [Planner] 问题拆解: {queries}")
                return queries
            else:
                return [question] # 格式不对，回退

        except Exception as e:
            print(f"⚠️ [Planner] 拆解失败 (Error: {e})，回退到原始问题")
            return [question]

def retrieve_multi_hop_context(question):
    """
    阶段一核心逻辑：Planner 分解 -> Retriever 批量检索 -> Aggregator 汇总
    """
    # 1. 初始化规划器
    planner = DecompositionAgent(client)

    # 2. 获取搜索列表 (例如: ["史瓦西半径", "引力红移"])
    sub_queries = planner.decompose(question)

    # 3. 【重要】总是把原始问题也加进去，防止拆解丢失了整体语义
    if question not in sub_queries:
        sub_queries.append(question)

    # 4. 批量检索与去重
    # 使用字典来去重，Key是公式名
    collected_formulas = {}

    print(f"📚 [Retriever] 开始多路检索...")

    for query in sub_queries:

        # 注意：这里调用的是你 formula_retrieval.py 里的 match_formula
        results = formula_retrieval_engine.match_formula(query, top_k=1, threshold=0.8)


        for f in results:
            f_name = f['formula_name']
            if f_name not in collected_formulas:
                collected_formulas[f_name] = f
                print(f"   -> 命中公式: {f_name} (来源: '{query}')")

    # 5. 格式化输出 (Aggregator)
    if not collected_formulas:
        return "No specific external formula found. Please rely on your internal knowledge and standard physical laws."

    # 拼接所有公式的 Prompt
    context_text = ""
    for f in collected_formulas.values():
        context_text += formula_retrieval_engine.generate_instruction(f) + "\n" + "-"*20 + "\n"

    return context_text

def generate_pot_prompt(question):
    """
    生成 Program-of-Thought (PoT) 提示词
    """
    # 1. 检索公式
    knowledge_injection = retrieve_multi_hop_context(question)
    # 2. 构建 Prompt
    system_prompt = (
        "You are an expert computational astrophysicist solving problems step-by-step.\n"
        "CRITICAL DECISION PROTOCOL (Follow strictly or fail):\n"
        "1.Textual Reasoning: When stating physical laws, establishing formulas, doing algebraic manipulation, or analyzing the problem, "
        "you MUST use Natural Language ONLY. DO NOT write Python code for abstract algebra.\n"
        "2.Numerical Calculation: ONLY when you need to calculate, you MUST use Python Code.\n"
        "3.Rule of Thumb: If there are no concrete numbers or units to compute in the current step, DO NOT use Python.\n"
        r"""
        --- EXAMPLE ---
Q: There are 15 trees in the grove. Grove workers will plant trees in the grove today. After they are done, there will be 21 trees. How many trees did the grove workers plant today?
A: We know the initial number of trees is 15 and the final number is 21. To find the number of planted trees, we need to subtract the initial amount from the final amount. Let's calculate this.
```python
initial_trees = 15
final_trees = 21
planted_trees = final_trees - initial_trees
print(planted_trees)
```
The answer is $\boxed{6}$.
"""
        r"""
        --- EXAMPLE 2 ---
Q: If there are 3 cars in the parking lot and 2 more cars arrive, how many cars are in the parking lot?
A: The initial number of cars is 3. The number of arriving cars is 2. We need to add them together to find the total.
```python
initial_cars = 3
arriving_cars = 2
total_cars = initial_cars + arriving_cars
print(total_cars)
```
The answer is $\boxed{5}$.
        """
    )
    user_prompt = (
        f"Problem: {question}\n"
        f"---------------------\n"
        f"Context & Formula Hints:\n{knowledge_injection}\n"
        f"---------------------\n"
        f"Please solve this problem step-by-step.\n"
        f"REQUIREMENTS:\n"
        f"1. In your Python code: Print the final result using `print(f'RESULT: {{result}}')`.\n"
        f"2. In your Text response: State the final answer clearly using LaTeX format $\\boxed{{answer}}$.\n"
    )

    print("==================================user_prompt===BEGIN=================================================")
    print(user_prompt)
    print("==================================user_prompt===END===================================================")

    return system_prompt, user_prompt

def generate_pot_prompt_without_formula(question):
    """
    生成 Program-of-Thought (PoT) 提示词
    """
    # 1. 检索公式
    # knowledge_injection = retrieve_multi_hop_context(question)
    # 2. 构建 Prompt
    system_prompt = (
        "You are an expert computational astrophysicist solving problems step-by-step.\n"
        "CRITICAL DECISION PROTOCOL (Follow strictly or fail):\n"
        "1.Textual Reasoning: When stating physical laws, establishing formulas, doing algebraic manipulation, or analyzing the problem, "
        "you MUST use Natural Language ONLY. DO NOT write Python code for abstract algebra.\n"
        "2.Numerical Calculation: ONLY when you need to calculate, you MUST use Python Code.\n"
        "3.Rule of Thumb: If there are no concrete numbers or units to compute in the current step, DO NOT use Python.\n"
        r"""
        --- EXAMPLE ---
Q: There are 15 trees in the grove. Grove workers will plant trees in the grove today. After they are done, there will be 21 trees. How many trees did the grove workers plant today?
A: We know the initial number of trees is 15 and the final number is 21. To find the number of planted trees, we need to subtract the initial amount from the final amount. Let's calculate this.
```python
initial_trees = 15
final_trees = 21
planted_trees = final_trees - initial_trees
print(planted_trees)
```
The answer is $\boxed{6}$.
"""
        r"""
        --- EXAMPLE 2 ---
Q: If there are 3 cars in the parking lot and 2 more cars arrive, how many cars are in the parking lot?
A: The initial number of cars is 3. The number of arriving cars is 2. We need to add them together to find the total.
```python
initial_cars = 3
arriving_cars = 2
total_cars = initial_cars + arriving_cars
print(total_cars)
```
The answer is $\boxed{5}$.
        """
    )
    user_prompt = (
        f"Problem: {question}\n"
        f"Please solve this problem step-by-step.\n"
        f"REQUIREMENTS:\n"
        f"1. In your Python code: Print the final result using `print(f'RESULT: {{result}}')`.\n"
        f"2. In your Text response: State the final answer clearly using LaTeX format $\\boxed{{answer}}$.\n"
    )

    print("====================================阶段二：推理=================================================")
    print(user_prompt)

    return system_prompt, user_prompt

def generate_prompt(question):
    """
    生成 Program-of-Thought (PoT) 提示词
    """
    # 1. 检索公式
    # formula_data, confidence = formula_retrieval_engine.match_formula(question)

    knowledge_injection = retrieve_multi_hop_context(question)
    # if formula_data and confidence > 0.3:
    #     knowledge_injection = formula_retrieval_engine.generate_instruction(formula_data)
    #     print(f"📚 [Knowledge] 注入公式: {formula_data['formula_name']} (Conf: {confidence:.2f})")

    # 2. 构建 Prompt
    system_prompt = ("You are an expert computational astrophysicist to solve the problem.\n")
    # system_prompt = (
    #     "You are an expert computational astrophysicist using a Jupyter Notebook style.\n"
    #     "Your Goal: Solve the problem step-by-step using a mix of Natural Language and Python Code.\n"
    #     "\n"
    #     "DECISION PROTOCOL (CRITICAL):\n"
    #     "1. **Conceptual Analysis**: If a step involves explaining concepts, choosing formulas, or logical deduction without numbers, use **Natural Language ONLY**.\n"
    #     "2. **Numerical Calculation**: If a step requires arithmetic, variable assignment, or unit conversion, you **MUST** use a Python code block.\n"
    #     "3. **Prohibited**: DO NOT calculate numbers manually in text. All numbers must come from code output.\n"
    #     "\n"
    #     "RULES for Code:\n"
    #     "1. Stateful Execution: Variables defined in previous blocks are available. Do not redefine imports.\n"
    #     "2. Attempt to use `astropy.units` and `astropy.constants` for rigorous physics.\n"
    #     "3. ⚠️ CRITICAL: If `astropy` causes errors or unit conversion issues, IMMEDIATELY switch to standard Python floats (using scientific notation) to ensure you get a numerical answer.\n"
    #     "4. Final Answer: At the very end, print the final result in SI units using `print(f'RESULT: ...')` within a code block."
    # )
    user_prompt = (
        f"Problem: {question}\n"
        f"---------------------\n"
        f"Context & Formula Hints:\n{knowledge_injection}\n"
        f"---------------------\n"
        f"Please state the final answer clearly using LaTeX format $\\boxed{{answer}}$.\n"
    )

    print("====================================阶段二：推理=================================================")
    print(user_prompt)

    return system_prompt,user_prompt

def extract_code_block(text):
    """提取 Markdown 中的 Python 代码块"""
    pattern = r"```python(.*?)```"
    matches = re.findall(pattern, text, re.DOTALL)
    if matches:
        return matches[-1].strip()  # 返回最后一个代码块
    return None

def clean_excel_string(value):
    """过滤掉 Excel 不支持的控制字符"""
    if not isinstance(value, str):
        return value
    # 匹配非法控制字符的正则表达式
    ILLEGAL_CHARACTERS_RE = re.compile(
        r"[\000-\010]|[\013-\014]|[\016-\037]"
    )
    return ILLEGAL_CHARACTERS_RE.sub(r"", value)

# 方案三：
def execute_hybrid_response(response_text, session):
    """
    解析回答中夹杂的所有代码块，并在同一个 session 中按顺序执行。

    Args:
        response_text (str): 模型生成的完整回复文本
        session (PythonSession): 当前问题的有状态会话

    Returns:
        tuple: (all_success: bool, execution_log: str, error_feedback: str)
    """
    # 提取所有 python 代码块 (支持 ```python 和 ```)
    code_blocks = re.findall(r"```(?:python)?(.*?)```", response_text, re.DOTALL)

    if not code_blocks:
        # 如果模型只输出了文本没有代码，视为一种特殊的“失败”（或者视情况允许纯文本）
        # 这里为了强制 PoT，我们返回 False 提示它写代码
        return False, "", "No Python code blocks found. Please write code to calculate."

    all_outputs = []

    print(f"    🔍 检测到 {len(code_blocks)} 个代码块")

    for i, code in enumerate(code_blocks):
        # 执行当前块 (状态会保留在 session 中)
        # success, output, feedback = session.execute(code)
        success, output, feedback = session.execute(code.strip())

        # 记录日志
        block_log = f"[Code Block {i+1}]:\n{code.strip()[:50]}..."
        exec_log = f"\n>>> Execution Output:\n{output.strip() if output else '(No Output)'}"

        if not success:
            # 一旦出错，立即中断
            error_msg = f"Error in Code Block {i+1}:\n{feedback}"
            print(f"    ❌ Block {i+1} Failed: {feedback}")
            return False, "\n".join(all_outputs), error_msg

        # 成功，记录输出
        all_outputs.append(f"{block_log}{exec_log}")
        print(f"    ✅ Block {i+1} Passed")

    return True, "\n".join(all_outputs), ""

def synthesize_final_report(client, question, history_content, execution_log):
    """
    阶段 2.5: 答案整理 (Answer Synthesis)
    将碎片化的 代码/日志/推理 整理成一份标准的解题报告
    """
    # print("===================================================================================")
    print(f"🚀 开始整理最终输出内容...")
    system_prompt = (
        "You are a scientific editor. Your task is to summarize the problem-solving process based on the provided execution logs.\n"
        "RULES:\n"
        "1. Strictly Fidelity: You must use the EXACT numerical results found in the [Execution Logs]. DO NOT recalculate.\n"
        "2. Please write the final structured solution report according to the following format:\n"
        "Given conditions: List the known information you obtained from the problem content here;\n"
        "Formulas used: List the formulas you will use here;\n"
        "Solution steps: Fill in the calculation process here, divided into Step 1, Step 2, Step 3... Step n (please try to simplify the content and number of steps);\n"
        "Final answer:Place the final result to the problem question here, with the final result represented using $\\boxed{{}}$. Please note that only one $\\boxed{{}}$ should appear.\n"
    )

    user_input = (
        f"Original Problem: {question}\n"
        f"========================================\n"
        f"Reasoning Process & Code Logs:\n"
        f"{history_content}\n"
        f"--- Execution Logs ---\n"
        f"{execution_log}\n"
    )
    print(f"user_input: {user_input[:50]}")

    try:
        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_input}
            ],
            temperature=0.1 # 低温，保证不胡编乱造
        )
        # print("===================================================================================")
        # print(f"🚀最终输出内容为：{response.choices[0].message.content}")
        # print("===================================================================================")
        return response.choices[0].message.content
    except Exception as e:
        return f"Error during synthesis: {e}\n\nRaw Logs:\n{execution_log}"

####################################方案三###################################
def iterative_reasoning_engine(system_prompt, user_prompt, max_retries=3):
    """
    基于‘交互式笔记本模式’的迭代推理引擎
    支持：混合文本/代码 + 状态保持 + 多轮反思 + 自适应纯文本降级兜底
    """
    # 1. 初始化
    # system_prompt, user_prompt = generate_pot_prompt(question)

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt}
    ]

    # 创建一个新的带记忆的会话 (Session)
    session = PythonSession()

    print(f"\n🚀 [Start] 问题: {question[:50]}...")

    for attempt in range(max_retries + 1):
        print(f"🤖 [Gen] 第 {attempt + 1} 轮推理中 (Attempt {attempt + 1}/{max_retries + 1})...")

        # 2. 模型生成
        try:
            response = client.chat.completions.create(
                model="deepseek-chat",
                messages=messages,
                temperature=0.3  # 保持低温以稳定生成代码
            )
            content = response.choices[0].message.content
            messages.append({"role": "assistant", "content": content})

            # 3. 混合执行 (Interactive Execution)
            print("⚙️ [Exec] 正在顺序执行代码块...")
            all_success, exec_log, error_msg = execute_hybrid_response(content, session)

            # 4. 结果判断与反馈环路
            if all_success:
                print(f"🎉 [Success] 代码执行或纯文本推理完毕，正在整理最终报告...")

                final_report = synthesize_final_report(
                    client=client,
                    question=question,
                    history_content=content,  # 模型原本生成的含代码的文本
                    execution_log=exec_log  # 代码执行出来的结果数值
                )
                print(f"[Report] 报告生成完毕。")
                return final_report

            else:
                # 失败分支
                print(f"❌ [Reflect] 执行中断。进入自我修正...")

                if attempt < max_retries:
                    # 构造反思 Prompt
                    reflection_prompt = (
                        f"Execution stopped due to an error.\n"
                        f"{error_msg}\n"
                        f"Please analyze the logic, fix the code, and RE-WRITE the entire solution (Text + Fixed Code)."
                    )
                    messages.append({"role": "user", "content": reflection_prompt})

                    # 策略A (严格): 重置 session，要求模型重写全部代码，防止脏数据污染
                    print("    🔄 重置代码会话状态...")
                    session = PythonSession()
                else:
                    # ==========================================
                    # 核心新增：自适应计算降级 (Fallback) 兜底机制
                    # ==========================================
                    print("⚠️ [Fallback] 达到最大代码重试次数，触发纯文本降级回退 (Adaptive Fallback)...")

                    fallback_prompt = (
                        "You have reached the maximum number of code execution attempts. "
                        "The code still encounters errors and cannot be executed successfully. "
                        "Now, ABANDON the code execution entirely. Based on the physical context, the formulas provided, "
                        "and your own reasoning, please output your best estimated final answer using NATURAL LANGUAGE ONLY. "
                        "Make sure to enclose your final numerical answer in \\boxed{}."
                    )
                    messages.append({"role": "user", "content": fallback_prompt})

                    try:
                        final_fallback_response = client.chat.completions.create(
                            model="deepseek-chat",
                            messages=messages,
                            temperature=0.3
                        ).choices[0].message.content

                        print("✅ [Fallback] 纯文本降级回答生成完毕。")
                        return final_fallback_response

                    except Exception as fallback_e:
                        print(f"⚠️ [Fallback Error] 纯文本降级 API 调用失败: {fallback_e}")
                        return f"{content}\n\n[Execution Failed & Fallback Failed]:\n{error_msg}"

        except Exception as e:
            print(f"⚠️ [System Error] API 调用或系统错误: {e}")
            return f"System Error: {str(e)}"

    return "Failed to generate valid solution after retries."

def reasoning_engine(question):
    """
    基于‘交互式笔记本模式’的迭代推理引擎
    支持：混合文本/代码 + 状态保持 + 多轮反思
    """
    # 1. 初始化
    # 注意：这里调用的是新的 generate_hybrid_prompt (需确保你已更新 Prompt 生成函数)
    system_prompt, user_prompt = generate_prompt(question)

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt}
    ]
    print(f"\n🚀 [Start] 问题: {question[:50]}...")

    # 2. 模型生成
    try:
        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=messages,
            temperature=0.3  # 保持低温以稳定生成代码
        )
        content = response.choices[0].message.content
        messages.append({"role": "assistant", "content": content})

        final_report = synthesize_final_report(
            client = client,
            question = question,
            history_content = content,
            execution_log=''# 模型原本生成的含代码的文本
        )
        print(f"[Report] 报告生成完毕。")
        return final_report



    except Exception as e:
        print(f"⚠️ [System Error] API 调用或系统错误: {e}")
        return f"System Error: {str(e)}"

# --- 主程序入口 ---
if __name__ == "__main__":
    SAMPLE_N = 3
    # 测试文件路径
    file_path = r"D:\PythonProjects\eval\Astro_Cal_Eval\eval_data\eval_file\Astro_Cal_Eval_EN_test.xlsx"
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # 写入表头
    sheet.cell(row=1, column=5, value="output")

    for i, row in enumerate(sheet.iter_rows(min_row=2)):
        question = row[2].value
        if not question: continue

        print(f"\n{'=' * 50}\nProcessing Row {i + 2}")

        # # 调用新的迭代推理引擎
        # final_response = iterative_reasoning_engine(question.strip())
        # # print(f'模型回答：{final_response}')
        # # 保存结果
        # # sheet.cell(row=i + 2, column=5, value=final_response)
        # clean_response = clean_excel_string(final_response)
        # sheet.cell(row=i + 2, column=5, value=clean_response)
        # wb.save(r"D:\PythonProjects\AstroCalcBoost2\main\new2\test_results\NASA_Eval_3_9_GPT-4o-mini_CoT+py_new2.xlsx")
        # 2. 并行采样
        print("🔍 正在进行问题拆解与多跳公式检索...")
        system_prompt, user_prompt = generate_pot_prompt(question.strip())

        all_candidate_completions = []
        print(f"🚀 启动并行推导 (n={SAMPLE_N}...")
        for id in range(SAMPLE_N):
            content = iterative_reasoning_engine(system_prompt,user_prompt,3)  # 阻塞：必须等这一行跑完
            all_candidate_completions.append([{"content": content}])
        #
        # 3. 奖励评分与 Index 索引
        if all_candidate_completions:
            # 调用你的 reward 函数计算每个回答的分数
            rewards = correctness_reward_func(all_candidate_completions, verbose=True)

            # 【关键索引步骤】：找到最高分对应的索引
            best_index = rewards.index(max(rewards))

            # 提取原始回答
            final_best_result = all_candidate_completions[best_index][0]["content"]

            print("\n" + "=" * 50)
            print(f"🏆 多数投票/评分胜出回答 (Index: {best_index})")
            print("=" * 50 + "\n")
            print(final_best_result)
            sheet.cell(row=i + 2, column=5, value=final_best_result)  # 第十列的索引是11（因为索引是从1开始的）
        else:
            print("未获取到有效回答。")
            sheet.cell(row=i + 2, column=5, value="未获取到有效回答")
        wb.save(r"D:\PythonProjects\AstroReflect\eval_results\deepseek0515_3复现2.xlsx")
        print(f'第{i}行已回答完毕！')
        print("\n" + "-" * 100)